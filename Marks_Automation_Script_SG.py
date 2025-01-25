# Sergio Garcia
# 12/11/2024

import argparse
import csv
import os
import ffmpeg
from pymongo import MongoClient
from datetime import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage

##############################################################################################################################

# These are parse arugments

def parse_args():
    parser = argparse.ArgumentParser(description="Process video segments and generate reports.")
    parser.add_argument("--process", type=str, help="Path to the MP4 video file", required=True)
    parser.add_argument("--baselight", type=str, help="Path to the Baselight file", required=True)
    parser.add_argument("--xytech", type=str, help="Path to the Xytech file", required=True)
    parser.add_argument("--outputXLS", type=str, help="Path to the output Excel file", required=True)
    parser.add_argument("--outputCSV", type=str, help="Path to the output CSV file", required=True)
    parser.add_argument("--outputDir", type=str, help="Directory to save extracted video segments", required=True)
    parser.add_argument("--thumbnailDir", type=str, help="Directory to save thumbnails", required=True)
    
    return parser.parse_args()

##############################################################################################################################

# These methods deal with MongoDB 

def connect_to_mongo():
    client = MongoClient('mongodb://localhost:27017/')
    mydb = client['The Crucible'] 
    return mydb

def insert_baselight_data(mydb, baselight_data):
    baselight_collection = mydb['baselight']
    for data in baselight_data:
        baselight_collection.insert_one(data)

def insert_xytech_data(mydb, xytech_data):
    xytech_collection = mydb['xytech']
    for data in xytech_data:
        xytech_collection.insert_one(data)

##############################################################################################################################

# These are from Project 1

def parse_baselight_file(file_path):
    baselight_locations = {}
    with open(file_path, 'r') as file:
        for line in file:
            line_parts = line.strip().split()
            if line_parts:
                location = line_parts[0]
                if location not in baselight_locations:
                    baselight_locations[location] = []
                baselight_locations[location].extend(line_parts[1:])
    return baselight_locations

def parse_xytech_file(file_path):
    xytech_info = {
        'producer': None,
        'operator': None,
        'job': None,
        'notes': None,
        'locations': []
    }

    with open(file_path, 'r') as file:
        is_notes_section = False
        notes_lines = []

        for line in file:
            line = line.strip()
            if line.startswith('Notes:'):
                is_notes_section = True
            elif is_notes_section and line:
                notes_lines.append(line)
            if not line:
                is_notes_section = False
            if line.startswith('/'):
                xytech_info['locations'].append(line)
            elif 'Producer:' in line:
                xytech_info['producer'] = line.split(': ')[1].strip()
            elif 'Operator:' in line:
                xytech_info['operator'] = line.split(': ')[1].strip()
            elif 'Job:' in line:
                xytech_info['job'] = line.split(': ')[1].strip()
            elif 'Notes:' in line:
                notes_split = line.split(': ')
                if len(notes_split) > 1:
                    xytech_info['notes'] = notes_split[1].strip()
    xytech_info['full_notes'] = ' '.join(notes_lines)
    return xytech_info

def combine_baselight_to_xytech(baselight_locations, xytech_info):
    combined_data = [['Producer', 'Operator', 'Job', 'Notes'],
                     [xytech_info['producer'], xytech_info['operator'], xytech_info['job'], xytech_info['notes']],
                     ['', '', '', ''],
                     ['Location', 'Frames to fix', '']]

    frame_to_locations = {}
    for xytech_loc in xytech_info['locations']:
        xytech_loc_suffix = xytech_loc.split('/production')[-1]
        for location, frames in baselight_locations.items():
            if xytech_loc_suffix in location:
                for frame in frames:
                    frame_to_locations.setdefault(frame, []).append(xytech_loc)

    frame_ranges = []
    current_location = None
    start_frame = end_frame = None
    sorted_frames = sorted(frame_to_locations.keys(), key=int)
    for i, frame in enumerate(sorted_frames):
        locations = frame_to_locations[frame]
        locations_str = ', '.join(locations)
        if locations_str != current_location:
            if start_frame is not None:
                frame_ranges.append((current_location, f"{start_frame}-{end_frame}" if start_frame != end_frame else str(start_frame)))
            current_location = locations_str
            start_frame = end_frame = int(frame)
        else:
            if int(frame) == end_frame + 1:
                end_frame = int(frame)
            else:
                frame_ranges.append((current_location, f"{start_frame}-{end_frame}" if start_frame != end_frame else str(start_frame)))
                start_frame = end_frame = int(frame)
    if start_frame is not None:
        frame_ranges.append((current_location, f"{start_frame}-{end_frame}" if start_frame != end_frame else str(start_frame)))

    for location, frame_range in frame_ranges:
        combined_data.append([location, frame_range, ""])

    return combined_data

##############################################################################################################################

# These methods deal with video processing

def get_video_duration(video_file):
    try:
        video_info = ffmpeg.probe(video_file)
        duration = float(video_info['format']['duration'])
        return duration
    except Exception as e:
        print(f"Error fetching video info: {e}")
        return 0

# Find matching frame ranges based on video duration
def find_matching_ranges(video_duration, database):
    baselight_collection = database['baselight']
    matched_ranges = []

    for entry in baselight_collection.find():
        location = entry["Location"]
        frames = entry["Frames"]
        for frame in frames:
            if int(frame) <= video_duration * 24:
                matched_ranges.append((location, frame))

    return matched_ranges

# Convert frame to timecode (hh:mm:ss.ms format)
def convert_frame_to_timecode(frame, fps=24):
    hours = int(frame / (fps * 3600))
    minutes = int((frame % (fps * 3600)) / (fps * 60))
    seconds = int((frame % (fps * 60)) / fps)
    milliseconds = int(((frame % fps) / fps) * 1000)
    return f"{hours:02}:{minutes:02}:{seconds:02}.{milliseconds:03}"

# Extract timecode ranges from video
def extract_timecode_ranges(video_file, frame_ranges, output_directory, fps=24):
    os.makedirs(output_directory, exist_ok=True)
    
    for i, (start_frame, end_frame) in enumerate(frame_ranges):
        start_time = convert_frame_to_timecode(start_frame, fps)
        end_time = convert_frame_to_timecode(end_frame, fps)
        output_path = os.path.join(output_directory, f"clip_{i + 1:03d}.mp4")

        try:
            print(f"Creating clip from {start_time} to {end_time} -> {output_path}")
            (
                ffmpeg
                .input(video_file, ss=start_time, to=end_time)
                .output(output_path, vcodec='libx264', acodec='aac', strict='experimental')
                .run(quiet=True, overwrite_output=True)
            )
        except Exception as e:
            print(f"Error creating clip for range {start_time} to {end_time}: {e}")

##############################################################################################################################

# These deal with excel

# Parse frame ranges from an Excel file
def parse_frame_ranges_from_xls(xls_file):
    workbook = load_workbook(xls_file)
    sheet = workbook.active
    frame_ranges = []

    for row in sheet.iter_rows(min_row=6):
        frame_range = row[1].value
        if frame_range:
            if "-" in frame_range:
                start_frame, end_frame = map(int, frame_range.split("-"))
            else:
                start_frame = end_frame = int(frame_range)
            frame_ranges.append((start_frame, end_frame))

    return frame_ranges

# Export matching ranges data to Excel
def export_excel(matching_ranges, output_file, thumbnail_folder, producer, operator, job):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Matching Ranges"

    sheet.append(["Producer", producer])
    sheet.append(["Operator", operator])
    sheet.append(["Job", job])
    sheet.append([]) 
    sheet.append(["Location", "Frame Range", "Timecode Range", "Thumbnail"])

    for entry in matching_ranges:
        location, frame_group = entry[:2]
        start_frame, end_frame = map(int, frame_group.split("-")) if "-" in frame_group else (int(frame_group), int(frame_group))

        start_timecode = convert_frame_to_timecode(start_frame)
        end_timecode = convert_frame_to_timecode(end_frame)
        timecode_range = f"{start_timecode} - {end_timecode}"

        row = [location, frame_group, timecode_range]
        sheet.append(row)

        thumbnail_path = os.path.join(thumbnail_folder, f"frame_{start_frame:04d}.jpg")
        try:
            img = OpenpyxlImage(thumbnail_path)
            img.width = 100
            img.height = 75
            sheet.add_image(img, f"D{sheet.max_row}")
        except Exception as e:
            print(f"Error adding thumbnail for {thumbnail_path}: {e}")

    workbook.save(output_file)
    
##############################################################################################################################

# Export CSV
    
def unsued_frames(frame_ranges, uploaded_frames, output_csv): 
    
    with open(output_csv, 'w', newline='') as csvfile:
        fieldnames = ['Location', 'Frame']
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
        for location, frame_range in frame_ranges:
            frame_numbers = frame_range.split('-')
            if len(frame_numbers) > 1:
                start_frame, end_frame = map(int, frame_numbers)
                for frame in range(start_frame, end_frame + 1):
                    if frame not in uploaded_frames:
                        writer.writerow({'Location': location, 'Frame': frame})
            else:
                frame = int(frame_numbers[0])
                if frame not in uploaded_frames:
                    writer.writerow({'Location': location, 'Frame': frame})

##############################################################################################################################

# Main

def main():
    
    args = parse_args()
    mydb = connect_to_mongo()

    baselight_data = parse_baselight_file(args.baselight)
    xytech_data = parse_xytech_file(args.xytech)
    insert_baselight_data(mydb, baselight_data)
    insert_xytech_data(mydb, xytech_data)

    matching_ranges = find_matching_ranges(get_video_duration(args.process), mydb)
    unsued_frames(matching_ranges, [], args.outputCSV) 

    export_excel(matching_ranges, args.outputXLS, args.thumbnailDir, xytech_data['producer'], xytech_data['operator'], xytech_data['job'])
    extract_timecode_ranges(args.process, matching_ranges, args.outputDir)

if __name__ == "__main__":
    main()